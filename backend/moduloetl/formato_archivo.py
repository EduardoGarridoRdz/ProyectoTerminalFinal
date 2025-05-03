from django.http import HttpResponse
import json
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
from django.views.decorators.csrf import csrf_exempt
from io import BytesIO
from openpyxl.styles import Font
import pandas as pd

@csrf_exempt
def FormatoArchivo(request):    
   import os

# Directorio actual de trabajo (desde donde se ejecuta el script)
   try:
      Archivo = request.FILES['archivo']

      if Archivo.content_type == "text/csv":
         
         df = pd.read_csv(Archivo, keep_default_na=False)
         
         output = BytesIO()
         
         with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Datos', startrow=9, startcol=0)  

         output.seek(0)
         wb = load_workbook(output)
         ws = wb.active
         # Añadir el logo
         logo = Image(r'moduloetl\Unicaribe.png') 
         logo.height = 200
         logo.width = 360
         ws.add_image(logo, 'A1') 

         # Añadir encabezado al archivo excel
         ws['I2'] = "UNIVERSIDAD DEL CARIBE"
         ws['I3'] = "SECRETARÍA DE PLANEACIÓN Y DESARROLLO INSTITUCIONAL"
         ws['I4'] = "DEPARTAMENTO DE CONTROL Y EVALUACIÓN"
         ws['I5'] = "INFORME DE ACTIVIDADES"
         
         # Ajustar el tamaño de fuente y el estilo
         ws['I2'].font = Font(size=12, bold=True)
         ws['I3'].font = Font(size=12, bold=True)
         ws['I4'].font = Font(size=12, bold=True)
         ws['I5'].font = Font(size=12, bold=True)
         ws['I6'].font = Font(size=12, bold=True)

         # Ajustar la alineación del texto
         ws['I2'].alignment = ws['I2'].alignment.copy(horizontal='center')
         ws['I3'].alignment = ws['I3'].alignment.copy(horizontal='center')
         ws['I4'].alignment = ws['I4'].alignment.copy(horizontal='center')
         ws['I5'].alignment = ws['I5'].alignment.copy(horizontal='center')
         ws['I6'].alignment = ws['I6'].alignment.copy(horizontal='center')

         # Crear un buffer para guardar el archivo
         output_buffer = BytesIO()
         # Guardar el archivo en el buffer
         wb.save(output_buffer)
         output_buffer.seek(0)
        
         # Devolver el archivo como respuesta
         response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
         )
         wb.save(response)
         return response
         
      else:
         return HttpResponse(
            json.dumps({'status': 'error', 'message': 'Tipo de archivo no soportado'}),
            content_type='application/json',
            status=400
         )
      
   except Exception as e:
     print("Error al procesar el archivo:", str(e))
     return HttpResponse(
        json.dumps({'status': 'error', 'message': str(e)}),
        content_type='application/json',
        status=400
      ) 
